import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl as vb
import pandas as pd

from scripts.calculation import vehicleCalculate, divisionCalculate
from scripts.fill import fillDivision, fillInfo
from scripts.process import processData
from scripts.constants import template_route


class OpsTrackerApp:
    def __init__(self, root): 
        self.root = root    # Load the root variable into this object
        self.this_month_data = None
        self.last_month_reports = None

        self.root.title("Monthly Ops Tracker Report System")
        self.root.geometry("520x340")
        self.root.resizable(False,False)

        self.load_background()
        self.build_ui()


    def load_background(self):
        image = Image.open("./images/example.jpg") # Define the photo component and get the file from the main file to locate
        self.photo = ImageTk.PhotoImage(image) 
        bg_label = tk.Label(self.root, image=self.photo) # The component load on the root
        bg_label.place(x=0, y=0, relwidth=1, relheight=1)
        self.root.attributes("-alpha",0.95)


    def build_ui(self):

        self.frame_buttons = tk.Frame(self.root)
        self.frame_buttons.pack(pady=30)

        self.frame_label = tk.Frame(self.frame_buttons)
        self.frame_label.pack(pady=0)

        tk.Button(self.frame_buttons, text="Upload This Month Raw Datasets", command=self.upload_this_datasets,font=("Arial",10)).pack(side="left",padx=10,pady=0,expand=True,fill="x")
        tk.Button(self.frame_buttons, text="Upload Last Month Reports", command=self.upload_last_reports,font=("Arial",10)).pack(side="left",padx=10,pady=5,expand=True,fill="x")

        self.this_label = tk.Label(self.frame_label, text="No File Uploaded", fg="gray",font=("Arial",8))
        self.this_label.pack(side="left",padx=25,expand=True,fill="x")

        self.last_label = tk.Label(self.frame_label, text="No File Uploaded", fg="gray",font=("Arial",8))
        self.last_label.pack(side="left",padx=25,expand=True,fill="x")

        self.report_frame = tk.Frame(self.root)
        self.report_frame.pack(padx=5,pady=10)

        self.log_text = tk.Text(self.report_frame, height=5, width=30, wrap="word",font=("Arial",8))
        self.log_text.pack(side="left",padx=5,pady=5)


        self.proc_frame = tk.Frame(self.report_frame)
        self.proc_frame.pack(side="left",pady=15)

        self.easter_btn = tk.Label(self.report_frame,text="",fg="black",font=("Arial",12))
        self.easter_btn.place(x=1,y=0)
        self.easter_btn.bind("<Button-1>",lambda e: self.show_easter_egg())

        self.text_label = tk.Label(self.proc_frame,text="Choose The Last Month",font=("Arial",8))
        self.text_label.pack(pady=0)

        self.combobox = ttk.Combobox(self.proc_frame, values=[],font=("Arial",8))
        self.combobox.set("None")
        self.combobox.pack(pady=6)



        self.this_month_input = tk.StringVar()
        tk.Label(self.proc_frame, text="This Month Report Date, eg: June 23",font=("Arial",8)).pack()
        tk.Entry(self.proc_frame, textvariable=self.this_month_input).pack(pady=2)

        tk.Button(self.proc_frame, text="Get The Report", command=self.get_report,font=("Arial",8)).pack(pady=5)



    def log(self, msg):
        self.log_text.insert(tk.END, msg + "\n") # Local function to output the log 
        self.log_text.see(tk.END)


    def upload_this_datasets(self):
        path = filedialog.askopenfilename(title="Choose this month datasets", filetypes=[("CSV File", "*.csv")])
        if path:
            self.this_month_data = path
            self.this_label.config(text=f"This month: {path.split('/')[-1]}")


    def upload_last_reports(self):
        path = filedialog.askopenfilename(title="Choose last month reports", filetypes=[("Excel File", "*.xlsx")])
        if path:
            self.last_month_reports = path
            self.last_label.config(text=f"Last month: {path.split('/')[-1]}")


            wb = vb.load_workbook(path)
            sheets = wb.sheetnames
            self.combobox['values'] = sheets
            self.combobox.set("Select the month")
            for title in sheets:
                self.log(f"Sheet: {title}")


    def on_select(self, event):
        self.log(f"Selected: {self.combobox.get()}")


    def process_report(self):
        wb = vb.load_workbook(self.last_month_reports)
        ws_name = self.combobox.get()
        last_month = ws_name.split()[0].capitalize()
        ws = wb[ws_name]

        template = vb.load_workbook(template_route)
        tem = template.active
        tem.title = self.this_month_input.get()
        this_month = self.this_month_input.get().split()[0].capitalize()

        tem = fillInfo(ws, tem, this_month, last_month)

        test = pd.read_csv(self.this_month_data)
        final_table, Total_OOS, Total_7DayUtilization, Total_WeekdayUtilization, Total_MilesDriven, Total_FobbedIn = divisionCalculate(test)
        final_result_Division_vehicle, final_result_vehicle = vehicleCalculate(test)
        tem = fillDivision(tem, final_table, Total_OOS, Total_7DayUtilization, Total_WeekdayUtilization, Total_MilesDriven, Total_FobbedIn, final_result_Division_vehicle, final_result_vehicle)

        return template


    def get_report(self):
        if not self.this_month_data or not self.last_month_reports:
            messagebox.showwarning("Missing File", "Please upload both datasets and reports")
            return
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx")],
            title="Save",
            initialfile=f"{self.this_month_input.get().split()[0].capitalize()} Monthly Ops Tracker Report"
        )
        if save_path:
            report = self.process_report()
            report.save(save_path)
            messagebox.showinfo("Success", f"Report saved as {save_path}")
        else:
            messagebox.showinfo("Cancelled", "Save cancelled.")

    def show_easter_egg(self):
        messagebox.showinfo("Surprise!!!","Made by Ruiyu(Ricardo) Yan Summer 2025")



    